from django.core.management.base import BaseCommand
from django.conf import settings
from gym.models import Member, Payment
import pandas as pd
import os
from datetime import datetime
from datetime import datetime, time
from django.utils.timezone import localtime
import openpyxl
from openpyxl.utils import get_column_letter


class Command(BaseCommand):
    help = "Backup Gym Data to Excel"

    def handle(self, *args, **kwargs):

        # ---------------- MEMBERS ---------------- #
        members_data = []

        for m in Member.objects.all():
            members_data.append({
                "Member Name": m.name,
                "Phone": m.phone,
                "Join Date": pd.to_datetime(m.join_date) if m.join_date else None,
            })

        members_df = pd.DataFrame(members_data)
        members_df.insert(0, "S.No", range(1, len(members_df) + 1))

        if not members_df.empty:
            members_df["Join Date"] = pd.to_datetime(members_df["Join Date"])


        # ---------------- PAYMENTS ---------------- #
        payments_data = []

        payments = Payment.objects.select_related(
            'subscription',
            'subscription__member'
        ).all()

        for p in payments:
            payments_data.append({
                "Member Name": p.subscription.member.name,
                "Amount Paid": float(p.amount_paid),
                "Payment Method": p.payment_method,
                "Payment Date": pd.to_datetime(p.payment_date),
            })

        payments_df = pd.DataFrame(payments_data)
        payments_df.insert(0, "S.No", range(1, len(payments_df) + 1))

        if not payments_df.empty:
            # keep as pandas datetime64 (not python date) so ExcelWriter
            # and openpyxl can apply the `datetime_format`/`date_format`
            payments_df["Payment Date"] = pd.to_datetime(payments_df["Payment Date"]) 


        # ---------------- FILE LOCATION ---------------- #
        from django.conf import settings
        import os

        now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

        backup_dir = os.path.join(settings.BASE_DIR, "backups")

# create backups folder if missing
        os.makedirs(backup_dir, exist_ok=True)

        filename = os.path.join(backup_dir, f"gym_backup_{now}.xlsx")


        # ---------------- WRITE EXCEL ---------------- #
        with pd.ExcelWriter(
               filename,
               engine="openpyxl",
               # Use lowercase Excel date format tokens so months display correctly
               date_format="dd-mm-yyyy",
               datetime_format="dd-mm-yyyy"
              ) as writer:


            members_df.to_excel(writer, sheet_name="Members", index=False)
            payments_df.to_excel(writer, sheet_name="Payments", index=False)


        self.stdout.write(self.style.SUCCESS("Backup Created Successfully"))

        # ---------------- FIX EXCEL CELL FORMATS ----------------
        # openpyxl tends to respect cell number formats better than relying
        # on pandas' date_format strings. Explicitly set date formats for the
        # 'Join Date' and 'Payment Date' columns so months display correctly.
        wb = openpyxl.load_workbook(filename)

        def set_date_column_format(sheet_name, header_name, fmt='dd-mm-yyyy'):
            if sheet_name not in wb.sheetnames:
                return
            ws = wb[sheet_name]
            # find header column
            col_idx = None
            for cell in ws[1]:
                if (cell.value or '') == header_name:
                    col_idx = cell.column
                    break

            if not col_idx:
                return

            col_letter = get_column_letter(col_idx)
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                if cell.value is None:
                    continue
                # Ensure value is a proper date/datetime for Excel
                # openpyxl treats Python date/datetime types specially.
                # If it's a string, try to parse ISO format to datetime.
                try:
                    val = cell.value
                    if isinstance(val, str):
                        try:
                            parsed = datetime.fromisoformat(val)
                            cell.value = parsed
                        except Exception:
                            # leave as-is if parsing fails
                            pass
                except Exception:
                    pass

                # set number format for date/time cells
                cell.number_format = fmt

        set_date_column_format('Members', 'Join Date', fmt='dd-mm-yyyy')
        set_date_column_format('Payments', 'Payment Date', fmt='dd-mm-yyyy')

        wb.save(filename)
